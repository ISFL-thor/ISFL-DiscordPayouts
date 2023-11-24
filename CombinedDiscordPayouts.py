import requests
import csv
import datetime
import openpyxl
import os

EXCEL_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Usernames.xlsx")


def get_username_mapping():
    # Load Excel file and map Discord usernames to desired usernames.
    # Returns a dictionary of the mapping.

    print("Loading username mappings from Excel...")
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = workbook.active
        username_mapping = {}

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=2, values_only=True):
            discord_username, desired_username = row
            username_mapping[str(discord_username).lower()] = desired_username

        print("Username mappings loaded successfully!")
        return username_mapping
    except Exception as e:
        print(f"Error reading the excel file: {e}")
        return {}


def get_mee6_leaderboard(SERVER_ID):
    # Fetch the Mee6 leaderboard data.
    # Returns a list of players.

    MEE6_API_URL = f"https://mee6.xyz/api/plugins/levels/leaderboard/{SERVER_ID}"
    print("Fetching data from Mee6 API...")
    try:
        response = requests.get(MEE6_API_URL)
        response.raise_for_status()
        print("Data fetched successfully from Mee6 API!")
        return response.json().get("players", [])
    except requests.RequestException as e:
        print(f"Failed to fetch data from Mee6 API: {e}")
        return []


def save_to_csv(users, mapping, prefix):
    # Save the user leaderboard data to a CSV file.
    # Returns a list of unmatched usernames.

    print("Saving data to CSV...")
    current_time = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"{prefix}Users_{current_time}.csv"
    unmatched_usernames = []

    try:
        with open(filename, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            for user in users:
                username_lower = user["username"].lower()
                desired_name = mapping.get(username_lower, user["username"])
                if username_lower not in mapping:
                    unmatched_usernames.append(user["username"])
                writer.writerow([desired_name, user["level"] * 100000])
        print(f"Data saved successfully to {filename}")
        return unmatched_usernames
    except Exception as e:
        print(f"Error saving data to CSV: {e}")
        return unmatched_usernames


def save_unmatched_to_excel(unmatched_usernames):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.title = "New Names"
    sheet['A1'] = 'Unmatched Usernames'

    for index, username in enumerate(unmatched_usernames, 2):
        sheet[f'A{index}'] = username

    file_name = "NEWNAMES.xlsx"
    workbook.save(file_name)
    return file_name


def main():
    SERVER_IDS = ["317388657994760194", "360525334472556544"]
    PREFIXES = ["ISFL", "DSFL"]
    all_unmatched_usernames = []

    mapping = get_username_mapping()

    for SERVER_ID, prefix in zip(SERVER_IDS, PREFIXES):
        users = get_mee6_leaderboard(SERVER_ID)

        if not users:
            print("No data to save.")
            continue

        unmatched = save_to_csv(users, mapping, prefix)
        all_unmatched_usernames.extend(unmatched)

    if all_unmatched_usernames:
        file_name = save_unmatched_to_excel(all_unmatched_usernames)
        print(f"\nUsernames with no match in Usernames.xlsx have been saved to {file_name}.")
        for username in all_unmatched_usernames:
            print(username)
    else:
        print("\nAll usernames have a match in Usernames.xlsx")

    input("Press Enter to exit...")

if __name__ == "__main__":
    main()